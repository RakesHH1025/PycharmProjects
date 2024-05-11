from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook('example.xlsx')

# Select the active worksheet (by default, it's the first sheet)
worksheet = workbook.active

# Iterate through rows and print data
for row in worksheet.iter_rows(values_only=True):
    print(row)
